import copy
import openpyxl
from openpyxl.utils import get_column_letter
import shutil
import PIL


def copyExcel(originFileName, newFileName, replaceDict, sheetName):
    # 复制源文件
    shutil.copy(originFileName, newFileName)

    # 替换文件
    wb = openpyxl.load_workbook(newFileName)

    for sheetname in wb.sheetnames:        
        sheet = wb[sheetname]
        if sheetname != sheetName:
            wb.remove(sheet)
            continue

        for row in sheet.iter_rows():
            for cell in row:
                # 查找是否有关键字,并替换内容
                if cell.value in replaceDict.keys():
                    cell.value = replaceDict[cell.value]

    wb.save(newFileName)
    wb.close()


def getFromExcel(fileName,worksheet):
    """返回两个Dict:  itemDict   detailDict"""
    # itemDict:从第二行开始获取关键字的子选项,返回字典
    workBook = openpyxl.load_workbook(fileName)
    if worksheet in workBook.sheetnames:
        workSheet = workBook[worksheet]

    itemDict = {}
    for row in range(2, workSheet.max_row + 1):
        if not workSheet['A'+str(row)].value: # 去除为None的情况
            continue
        if workSheet['A'+str(row)].value not in itemDict:
            itemDict[workSheet['A'+str(row)].value] = [workSheet['B'+str(row)].value]
        else:
            itemDict[workSheet['A'+str(row)].value].append(workSheet['B'+str(row)].value)
    """
    itemDict = {
    'XIANCAI': ['HDMI 线缆-5m', 'HDMI 线缆-10m', 'HDMI 线缆-15m'],
    'ZHIJIA': ['拼接屏前维护平推钣金支架单元', '拼接屏落地式支架(标准直立式)'],
    'ANZHUANGFEI': ['LCD拼接系...
    }
    """

    # detailDict生成
    """
    keyList = []
    for key in itemDict.keys():
        # 去除key为None的情况
        if key:  
            keyList = keyList + itemDict[key]
    # print(keyList)
        
    detailDict = {}
    for row in range(2, workSheet.max_row + 1):
        if workSheet['B'+str(row)].value in keyList:
            detailDict[workSheet['B'+str(row)].value] = {}
            for column in range(3,workSheet.max_column+1):
                # print(workSheet['A'+str(row)].value) # data.xlsx中的KEY列不能重名
                detailDict[workSheet['B'+str(row)].value][workSheet['A'+str(row)].value+'_'+workSheet[chr(column+64)+'1'].value] = workSheet[chr(column+64)+str(row)].value
    """
    detailDict = {}
    for row in range(2, workSheet.max_row + 1):
        detailDict[workSheet['B'+str(row)].value] = {}
        for column in range(3,workSheet.max_column+1):
            # data.xlsx中的KEY列不能重名
            if workSheet['A'+str(row)].value and workSheet['B'+str(row)].value: # 去除 ITEM 或者 KEY 为 None的情况
                detailDict[workSheet['B'+str(row)].value][workSheet['A'+str(row)].value+'_'+workSheet[chr(column+64)+'1'].value] = workSheet[chr(column+64)+str(row)].value
    """
     {'HDMI 线缆-5m': {'XIANCAI_XINGHAO': 'CAB-HI-05M', 'XIANCAI_MINGCHENG': 'HDMI 线缆-5m', 'XIANCAI_CANSHU': '说明:5米线缆：\n类型：
    HDMI线缆\n颜色：黑色\n光纤材质、无损传输、超长距离传输无信号衰减', 'XIANCAI_DANJIA': 30, 'XIANCAI_FENBIANLV': None, 'XIANCAI_CHICUN':
     None, 'XIANCAI_GONGLV': None, 'XIANCAI_LN-DH7508-S': None, 'XIANCAI_LN-DH7512-S': None, 'XIANCAI_LN-DH7516-S': None}, 'HDMI 线缆
    -10m': {'XIANCAI_XINGHAO': 'CAB-HI-10M', 'XIANCAI_MINGCHENG': 'HDMI 线缆-10m', 'XIANCAI_CANSHU': '说明:10米线缆：\n类型：HDMI线缆
    \n颜色：黑色\n光纤材质、无损传输、超长距离传输无信号衰减', 'XIANCAI_DANJIA': 65, 'XIANCAI_FENBIANLV': None, 'XIANCAI_CHICUN': None, 
    'XIANCAI_GONGLV': None, 'XIANCAI_LN-DH7508-S': None, 'XIANCAI_LN-DH7512-S': None, 'XIANCAI_LN-DH7516-S': None}, 'HDMI 线缆-15m': 
    {'XIANCAI_XINGHAO': 'CAB-HI-15M', 'XIANCAI_MINGCHENG': 'HDMI 线缆-15m', 'XIANCAI_CANSHU': '说明:15米线缆：\n类型：HDMI线缆\n颜色
    ：黑色\n光纤材质、无损传输、超长距离传输无信号衰减', 'XIANCAI_DANJIA': 110, 'XIANCAI_FENBIANLV': None, 'XIANCAI_CHICUN': None, 
    'XIANCAI_GONGLV': None, 'XIANCAI_LN-DH7508-S': None, 'XIANCAI_LN-DH7512-S': None, 'XIANCAI_LN-DH7516-S': None}, '拼接屏前维护平推
    钣金支架单元': {'ZHIJIA_XINGHAO': 'HB-3655-G-B', 'ZHIJIA_MINGCHENG': '拼接屏前维护平推钣金支架单元', 'ZHIJIA_CANSHU': '颜色：黑色；\n材料
    ：SPCC优
    """

    workBook.close()
    
    return itemDict,detailDict


def writeInCanShu(fileName,worksheet,swCanShu,columnName="SHANGWUCANSHU"):
    wb = openpyxl.load_workbook(fileName)
    ws = workSheet = wb[worksheet]

    key_c = -1
    for column in range(1,ws.max_column+1):
        if workSheet[chr(column+64)+'1'].value == columnName:
            key_c = column
            break
    if key_c != -1:
        for row in range(2, workSheet.max_row + 1):
            if (ws['B'+str(row)].value,ws['C'+str(row)].value) in swCanShu: # 没有这个参数就不写入
                ws[chr(key_c+64)+str(row)].value = swCanShu[(ws['B'+str(row)].value,ws['C'+str(row)].value)]
        
    wb.save(fileName)
    wb.close()

if __name__ == '__main__':
    k,d= getFromExcel('data.xlsx')
    print(k,'\n',d)
